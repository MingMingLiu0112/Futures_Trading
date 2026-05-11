#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回测结果导出模块

支持导出为 Excel 和 PDF 格式。

用法:
```python
from backtest.backtest_exporter import BacktestExporter

exporter = BacktestExporter(result_data)
excel_bytes = exporter.to_excel()
exporter.to_pdf('report.pdf')
```
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

# Excel支持
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENEXCEL_AVAILABLE = True
except ImportError:
    OPENEXCEL_AVAILABLE = False

# PDF支持
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class BacktestExporter:
    """
    回测结果导出器

    将回测结果导出为 Excel 或 PDF 格式。
    """

    def __init__(self, result_data: Dict[str, Any]):
        """
        Args:
            result_data: 回测结果字典，应包含:
                - strategy_name: 策略名
                - initial_balance: 初始资金
                - final_balance: 最终资金
                - statistics: 绩效指标字典
                - trades: 交易记录列表
                - equity_curve: 权益曲线
        """
        self.result_data = result_data

    def to_dict(self) -> Dict[str, Any]:
        """导出为字典格式"""
        stats = self.result_data.get('statistics', {})
        trades = self.result_data.get('trades', [])
        equity_curve = self.result_data.get('equity_curve', [])

        return {
            'summary': {
                'strategy_name': self.result_data.get('strategy_name', 'Unknown'),
                'initial_balance': self.result_data.get('initial_balance', 0),
                'final_balance': self.result_data.get('final_balance', 0),
                'total_return': stats.get('total_return', 0),
                'total_trades': stats.get('total_trades', 0),
                'win_rate': stats.get('win_rate', 0),
                'sharpe_ratio': stats.get('sharpe_ratio', 0),
                'sortino_ratio': stats.get('sortino_ratio', 0),
                'calmar_ratio': stats.get('calmar_ratio', 0),
                'max_drawdown': stats.get('max_drawdown', 0),
            },
            'statistics': stats,
            'trades': trades,
            'equity_curve': equity_curve,
            'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }

    def to_excel(self, filepath: str = None) -> Optional[bytes]:
        """
        导出为 Excel 文件

        Args:
            filepath: 保存路径，如果为 None 则返回 bytes

        Returns:
            Excel 文件 bytes 或 None
        """
        if not OPENEXCEL_AVAILABLE:
            raise ImportError("需要 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        stats = self.result_data.get('statistics', {})
        trades = self.result_data.get('trades', [])
        equity_curve = self.result_data.get('equity_curve', [])

        # ========== 汇总 sheet ==========
        ws_summary = wb.create_sheet('汇总')
        self._write_summary_sheet(ws_summary, stats)

        # ========== 统计 sheet ==========
        ws_stats = wb.create_sheet('绩效指标')
        self._write_stats_sheet(ws_stats, stats)

        # ========== 交易明细 sheet ==========
        ws_trades = wb.create_sheet('交易明细')
        self._write_trades_sheet(ws_trades, trades)

        # ========== 权益曲线 sheet ==========
        ws_equity = wb.create_sheet('权益曲线')
        self._write_equity_sheet(ws_equity, equity_curve)

        if filepath:
            wb.save(filepath)
            return None
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

    def to_pdf(self, filepath: str) -> bool:
        """
        导出为 PDF 文件

        Args:
            filepath: 保存路径

        Returns:
            是否成功
        """
        if not PDF_AVAILABLE:
            raise ImportError("需要 reportlab: pip install reportlab")

        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        stats = self.result_data.get('statistics', {})
        trades = self.result_data.get('trades', [])
        elements = []

        # 标题
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # 居中
        )
        elements.append(Paragraph('回测报告', title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # 中文支持
        try:
            pdfmetrics.registerFont(TTFont('SimHei', '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc'))
            chinese_font = 'SimHei'
        except Exception:
            chinese_font = 'Helvetica'

        summary_data = [
            ['策略名称', self.result_data.get('strategy_name', 'Unknown')],
            ['初始资金', f"¥{self.result_data.get('initial_balance', 0):,.2f}"],
            ['最终资金', f"¥{self.result_data.get('final_balance', 0):,.2f}"],
            ['总收益率', f"{stats.get('total_return', 0):.2f}%"],
            ['年化收益', f"{stats.get('annual_return', 0):.2f}%"],
            ['总交易次数', str(stats.get('total_trades', 0))],
            ['胜率', f"{stats.get('win_rate', 0):.2f}%"],
            ['夏普比率', f"{stats.get('sharpe_ratio', 0):.4f}"],
            ['索提诺比率', f"{stats.get('sortino_ratio', 0):.4f}"],
            ['卡尔玛比率', f"{stats.get('calmar_ratio', 0):.4f}"],
            ['最大回撤', f"¥{stats.get('max_drawdown', 0):,.2f}"],
            ['最大连续亏损', str(stats.get('max_consecutive_losses', 0))],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8E8E8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # 交易明细表（只显示前20条）
        if trades:
            trade_header = ['交易ID', '方向', '入场时间', '出场时间', '入场价', '出场价', '数量', '盈亏', '盈亏%', '出场原因']
            trade_rows = [trade_header]
            for t in trades[:20]:
                trade_rows.append([
                    t.get('trade_id', ''),
                    t.get('direction', ''),
                    t.get('entry_time', ''),
                    t.get('exit_time', ''),
                    f"{t.get('entry_price', 0):.2f}",
                    f"{t.get('exit_price', 0):.2f}",
                    str(t.get('quantity', 0)),
                    f"¥{t.get('pnl', 0):.2f}",
                    f"{t.get('pnl_pct', 0):.2f}%",
                    t.get('exit_reason', ''),
                ])

            trade_table = Table(trade_rows, repeatRows=1)
            trade_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(Paragraph('交易明细（前20条）', styles['Heading2']))
            elements.append(trade_table)

        doc.build(elements)
        return True

    def to_pdf_buffer(self, buffer) -> bool:
        """
        导出为 PDF 到内存缓冲区（用于API返回）

        Args:
            buffer: io.BytesIO 缓冲区

        Returns:
            是否成功
        """
        if not PDF_AVAILABLE:
            raise ImportError("需要 reportlab: pip install reportlab")

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        stats = self.result_data.get('statistics', {})
        trades = self.result_data.get('trades', [])
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1
        )
        elements.append(Paragraph('回测报告', title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # 中文支持
        try:
            pdfmetrics.registerFont(TTFont('SimHei', '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc'))
            chinese_font = 'SimHei'
        except Exception:
            chinese_font = 'Helvetica'

        summary_data = [
            ['策略名称', self.result_data.get('strategy_name', 'Unknown')],
            ['初始资金', f"¥{self.result_data.get('initial_balance', 0):,.2f}"],
            ['最终资金', f"¥{self.result_data.get('final_balance', 0):,.2f}"],
            ['总收益率', f"{stats.get('total_return', 0):.2f}%"],
            ['年化收益', f"{stats.get('annual_return', 0):.2f}%"],
            ['总交易次数', str(stats.get('total_trades', 0))],
            ['胜率', f"{stats.get('win_rate', 0):.2f}%"],
            ['夏普比率', f"{stats.get('sharpe_ratio', 0):.4f}"],
            ['索提诺比率', f"{stats.get('sortino_ratio', 0):.4f}"],
            ['卡尔玛比率', f"{stats.get('calmar_ratio', 0):.4f}"],
            ['最大回撤', f"¥{stats.get('max_drawdown', 0):,.2f}"],
            ['最大连续亏损', str(stats.get('max_consecutive_losses', 0))],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8E8E8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        if trades:
            trade_header = ['交易ID', '方向', '入场时间', '出场时间', '入场价', '出场价', '数量', '盈亏', '盈亏%', '出场原因']
            trade_rows = [trade_header]
            for t in trades[:20]:
                trade_rows.append([
                    str(t.get('trade_id', '')),
                    t.get('direction', ''),
                    t.get('entry_time', ''),
                    t.get('exit_time', ''),
                    f"{t.get('entry_price', 0):.2f}",
                    f"{t.get('exit_price', 0):.2f}",
                    str(t.get('quantity', 0)),
                    f"¥{t.get('pnl', 0):.2f}",
                    f"{t.get('pnl_pct', 0):.2f}%",
                    t.get('exit_reason', ''),
                ])

            trade_table = Table(trade_rows, repeatRows=1)
            trade_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), chinese_font),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(Paragraph('交易明细（前20条）', styles['Heading2']))
            elements.append(trade_table)

        doc.build(elements)
        return True

    def _write_summary_sheet(self, ws, stats: Dict):
        """写入汇总 sheet"""
        ws['A1'] = '回测汇总'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        rows = [
            ['策略名称', self.result_data.get('strategy_name', 'Unknown')],
            ['初始资金', self.result_data.get('initial_balance', 0)],
            ['最终资金', self.result_data.get('final_balance', 0)],
            ['总收益率', f"{stats.get('total_return', 0):.2f}%"],
            ['年化收益', f"{stats.get('annual_return', 0):.2f}%"],
            ['夏普比率', stats.get('sharpe_ratio', 0)],
            ['索提诺比率', stats.get('sortino_ratio', 0)],
            ['卡尔玛比率', stats.get('calmar_ratio', 0)],
            ['最大回撤', stats.get('max_drawdown', 0)],
            ['最大回撤%', f"{stats.get('max_drawdown_pct', 0):.2f}%"],
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        for i, (k, v) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)

        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_stats_sheet(self, ws, stats: Dict):
        """写入统计指标 sheet"""
        ws['A1'] = '绩效指标详情'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        categories = {
            '基础统计': ['total_trades', 'win_count', 'loss_count', 'win_rate',
                        'avg_win', 'avg_loss', 'profit_loss_ratio', 'total_pnl'],
            '收益指标': ['total_return', 'annual_return', 'daily_pnl', 'return_std'],
            '风险指标': ['max_drawdown', 'max_drawdown_pct', 'sharpe_ratio',
                        'sortino_ratio', 'calmar_ratio', 'return_drawdown_ratio'],
            '连续性指标': ['max_consecutive_losses', 'max_consecutive_wins'],
        }

        row = 3
        for category, keys in categories.items():
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=1).font = Font(bold=True, color='4472C4')
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='D9E1F2')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

            for key in keys:
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=stats.get(key, 0))
                row += 1
            row += 1

        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _write_trades_sheet(self, ws, trades: List[Dict]):
        """写入交易明细 sheet"""
        headers = ['交易ID', '方向', '入场时间', '出场时间', '入场价', '出场价',
                   '数量', '盈亏', '盈亏%', '出场原因']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')

        for row_idx, trade in enumerate(trades, start=2):
            ws.cell(row=row_idx, column=1, value=trade.get('trade_id', ''))
            ws.cell(row=row_idx, column=2, value=trade.get('direction', ''))
            ws.cell(row=row_idx, column=3, value=trade.get('entry_time', ''))
            ws.cell(row=row_idx, column=4, value=trade.get('exit_time', ''))
            ws.cell(row=row_idx, column=5, value=trade.get('entry_price', 0))
            ws.cell(row=row_idx, column=6, value=trade.get('exit_price', 0))
            ws.cell(row=row_idx, column=7, value=trade.get('quantity', 0))
            ws.cell(row=row_idx, column=8, value=trade.get('pnl', 0))
            ws.cell(row=row_idx, column=9, value=trade.get('pnl_pct', 0))
            ws.cell(row=row_idx, column=10, value=trade.get('exit_reason', ''))

        widths = [12, 8, 20, 20, 12, 12, 8, 12, 10, 15]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_equity_sheet(self, ws, equity_curve: List[Dict]):
        """写入权益曲线 sheet"""
        headers = ['时间', '余额']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')

        for row_idx, eq in enumerate(equity_curve, start=2):
            ws.cell(row=row_idx, column=1, value=eq.get('time', ''))
            ws.cell(row=row_idx, column=2, value=eq.get('balance', 0))

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
