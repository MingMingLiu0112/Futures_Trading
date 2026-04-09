"""
期权链数据Excel导出Flask应用 - 完整版
支持完整期权链数据、PCR数据、希腊字母、历史数据导出
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
import json
import io
from typing import Dict, List, Optional, Tuple
import logging

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB 最大上传
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['EXPORT_FOLDER'] = 'exports'

# 创建必要的目录
for folder in [app.config['UPLOAD_FOLDER'], app.config['EXPORT_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

class OptionChainExporter:
    """期权链数据导出器"""
    
    def __init__(self):
        self.sample_data = self._generate_sample_data()
    
    def _generate_sample_data(self) -> Dict:
        """生成示例期权链数据"""
        # 生成示例数据
        current_price = 100.0
        strike_prices = np.arange(80, 121, 5)
        expiration_dates = [
            (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
            (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d'),
            (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')
        ]
        
        option_chain = []
        for expiry in expiration_dates:
            for strike in strike_prices:
                # 生成看涨期权数据
                call_data = {
                    'symbol': f'C_{strike}_{expiry.replace("-", "")}',
                    'type': 'call',
                    'strike': strike,
                    'expiry': expiry,
                    'last_price': round(np.random.uniform(0.5, 20.0), 2),
                    'bid': round(np.random.uniform(0.4, 19.0), 2),
                    'ask': round(np.random.uniform(0.6, 21.0), 2),
                    'volume': np.random.randint(100, 10000),
                    'open_interest': np.random.randint(1000, 50000),
                    'implied_volatility': round(np.random.uniform(0.15, 0.45), 4),
                    'delta': round(np.random.uniform(0.1, 0.9), 4),
                    'gamma': round(np.random.uniform(0.01, 0.05), 4),
                    'theta': round(np.random.uniform(-0.05, -0.01), 4),
                    'vega': round(np.random.uniform(0.01, 0.03), 4),
                    'rho': round(np.random.uniform(0.01, 0.02), 4),
                    'moneyness': 'ITM' if strike < current_price else ('ATM' if strike == current_price else 'OTM')
                }
                
                # 生成看跌期权数据
                put_data = {
                    'symbol': f'P_{strike}_{expiry.replace("-", "")}',
                    'type': 'put',
                    'strike': strike,
                    'expiry': expiry,
                    'last_price': round(np.random.uniform(0.5, 20.0), 2),
                    'bid': round(np.random.uniform(0.4, 19.0), 2),
                    'ask': round(np.random.uniform(0.6, 21.0), 2),
                    'volume': np.random.randint(100, 10000),
                    'open_interest': np.random.randint(1000, 50000),
                    'implied_volatility': round(np.random.uniform(0.15, 0.45), 4),
                    'delta': round(np.random.uniform(-0.9, -0.1), 4),
                    'gamma': round(np.random.uniform(0.01, 0.05), 4),
                    'theta': round(np.random.uniform(-0.05, -0.01), 4),
                    'vega': round(np.random.uniform(0.01, 0.03), 4),
                    'rho': round(np.random.uniform(-0.02, -0.01), 4),
                    'moneyness': 'ITM' if strike > current_price else ('ATM' if strike == current_price else 'OTM')
                }
                
                option_chain.extend([call_data, put_data])
        
        # 计算PCR数据
        pcr_data = []
        for expiry in expiration_dates:
            expiry_options = [opt for opt in option_chain if opt['expiry'] == expiry]
            calls = [opt for opt in expiry_options if opt['type'] == 'call']
            puts = [opt for opt in expiry_options if opt['type'] == 'put']
            
            pcr_volume = sum(p['volume'] for p in puts) / sum(c['volume'] for c in calls) if sum(c['volume'] for c in calls) > 0 else 0
            pcr_oi = sum(p['open_interest'] for p in puts) / sum(c['open_interest'] for c in calls) if sum(c['open_interest'] for c in calls) > 0 else 0
            
            pcr_data.append({
                'expiry': expiry,
                'pcr_volume': round(pcr_volume, 4),
                'pcr_open_interest': round(pcr_oi, 4),
                'total_volume': sum(opt['volume'] for opt in expiry_options),
                'total_open_interest': sum(opt['open_interest'] for opt in expiry_options),
                'avg_iv_call': round(np.mean([c['implied_volatility'] for c in calls]), 4),
                'avg_iv_put': round(np.mean([p['implied_volatility'] for p in puts]), 4),
                'iv_skew': round(np.mean([p['implied_volatility'] for p in puts]) - np.mean([c['implied_volatility'] for c in calls]), 4)
            })
        
        # 生成历史数据
        history_days = 30
        history_data = []
        base_date = datetime.now() - timedelta(days=history_days)
        
        for i in range(history_days):
            date = base_date + timedelta(days=i)
            history_data.append({
                'date': date.strftime('%Y-%m-%d'),
                'underlying_price': round(current_price + np.random.uniform(-5, 5), 2),
                'total_volume': np.random.randint(100000, 500000),
                'total_open_interest': np.random.randint(1000000, 5000000),
                'avg_iv': round(np.random.uniform(0.2, 0.4), 4),
                'pcr_volume': round(np.random.uniform(0.8, 1.2), 4),
                'pcr_oi': round(np.random.uniform(0.9, 1.1), 4)
            })
        
        return {
            'underlying': {
                'symbol': 'AAPL',
                'current_price': current_price,
                'change': 1.5,
                'change_percent': 1.52,
                'volume': 4567890,
                'market_cap': '2.8T'
            },
            'option_chain': option_chain,
            'pcr_data': pcr_data,
            'history_data': history_data,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def export_to_excel(self, data: Dict, filename: str = None) -> str:
        """
        导出数据到Excel文件
        
        Args:
            data: 包含期权链数据的字典
            filename: 输出文件名（可选）
            
        Returns:
            文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'option_chain_export_{timestamp}.xlsx'
        
        filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
        
        # 创建Excel写入器
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 基础信息sheet
            underlying_df = pd.DataFrame([data['underlying']])
            underlying_df.to_excel(writer, sheet_name='基础信息', index=False)
            
            # 2. 期权链数据sheet
            option_df = pd.DataFrame(data['option_chain'])
            option_df.to_excel(writer, sheet_name='期权链数据', index=False)
            
            # 3. PCR数据sheet
            pcr_df = pd.DataFrame(data['pcr_data'])
            pcr_df.to_excel(writer, sheet_name='PCR数据', index=False)
            
            # 4. 历史数据sheet
            history_df = pd.DataFrame(data['history_data'])
            history_df.to_excel(writer, sheet_name='历史数据', index=False)
            
            # 5. 希腊字母汇总sheet
            greek_summary = self._create_greek_summary(option_df)
            greek_summary.to_excel(writer, sheet_name='希腊字母汇总', index=False)
            
            # 6. 波动率曲面sheet
            vol_surface = self._create_volatility_surface(option_df)
            vol_surface.to_excel(writer, sheet_name='波动率曲面', index=False)
        
        logger.info(f"Excel文件已生成: {filepath}")
        return filepath
    
    def _create_greek_summary(self, option_df: pd.DataFrame) -> pd.DataFrame:
        """创建希腊字母汇总表"""
        summary_data = []
        
        # 按类型和到期日分组
        for expiry in option_df['expiry'].unique():
            for option_type in ['call', 'put']:
                mask = (option_df['expiry'] == expiry) & (option_df['type'] == option_type)
                df_subset = option_df[mask]
                
                if len(df_subset) > 0:
                    summary_data.append({
                        '到期日': expiry,
                        '类型': '看涨' if option_type == 'call' else '看跌',
                        '数量': len(df_subset),
                        '平均Delta': round(df_subset['delta'].mean(), 4),
                        '平均Gamma': round(df_subset['gamma'].mean(), 4),
                        '平均Theta': round(df_subset['theta'].mean(), 4),
                        '平均Vega': round(df_subset['vega'].mean(), 4),
                        '平均Rho': round(df_subset['rho'].mean(), 4),
                        '平均IV': round(df_subset['implied_volatility'].mean(), 4),
                        'Delta范围': f"{round(df_subset['delta'].min(), 4)} - {round(df_subset['delta'].max(), 4)}",
                        'IV范围': f"{round(df_subset['implied_volatility'].min(), 4)} - {round(df_subset['implied_volatility'].max(), 4)}"
                    })
        
        return pd.DataFrame(summary_data)
    
    def _create_volatility_surface(self, option_df: pd.DataFrame) -> pd.DataFrame:
        """创建波动率曲面数据"""
        surface_data = []
        
        # 获取唯一的到期日和行权价
        expiries = sorted(option_df['expiry'].unique())
        strikes = sorted(option_df['strike'].unique())
        
        for expiry in expiries:
            for strike in strikes:
                # 查找该到期日和行权价的期权
                calls = option_df[(option_df['expiry'] == expiry) & 
                                 (option_df['strike'] == strike) & 
                                 (option_df['type'] == 'call')]
                puts = option_df[(option_df['expiry'] == expiry) & 
                                (option_df['strike'] == strike) & 
                                (option_df['type'] == 'put')]
                
                call_iv = calls['implied_volatility'].mean() if not calls.empty else None
                put_iv = puts['implied_volatility'].mean() if not puts.empty else None
                avg_iv = np.mean([iv for iv in [call_iv, put_iv] if iv is not None]) if any([call_iv, put_iv]) else None
                
                surface_data.append({
                    '到期日': expiry,
                    '行权价': strike,
                    '看涨IV': round(call_iv, 4) if call_iv else '',
                    '看跌IV': round(put_iv, 4) if put_iv else '',
                    '平均IV': round(avg_iv, 4) if avg_iv else '',
                    'IV偏度': round(put_iv - call_iv, 4) if call_iv and put_iv else ''
                })
        
        return pd.DataFrame(surface_data)
    
    def export_to_excel_with_formatting(self, data: Dict, filename: str = None) -> str:
        """
        导出数据到Excel文件（带格式化）
        
        Args:
            data: 包含期权链数据的字典
            filename: 输出文件名（可选）
            
        Returns:
            文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'option_chain_formatted_{timestamp}.xlsx'
            
            filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
            
            # 创建Workbook
            wb = Workbook()
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            data_font = Font(size=11)
            money_font = Font(size=11, color="0000FF")
            negative_font = Font(size=11, color="FF0000")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 基础信息sheet
            ws1 = wb.active
            ws1.title = "基础信息"
            
            # 写入标题
            ws1['A1'] = "标的资产信息"
            ws1['A1'].font = Font(bold=True, size=14, color="366092")
            ws1.merge_cells('A1:E1')
            
            # 写入表头
            headers = ['字段', '值']
            ws1.append(headers)
            
            for cell in ws1[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            underlying = data['underlying']
            rows = [
                ['标的代码', underlying['symbol']],
                ['当前价格', f"${underlying['current_price']:.2f}"],
                ['涨跌', f"{underlying['change']:.2f}"],
                ['涨跌幅', f"{underlying['change_percent']:.2f}%"],
                ['成交量', f"{underlying['volume']:,}"],
                ['市值', underlying['market_cap']],
                ['数据生成时间', data['generated_at']]
            ]
            
            for row in rows:
                ws1.append(row)
            
            # 调整列宽
            for col in range(1, 3):
                ws1.column_dimensions[get_column_letter(col)].width = 20
            
            # 2. 期权链数据sheet
            ws2 = wb.create_sheet("期权链数据")
            
            # 写入标题
            ws2['A1'] = "期权链详细数据"
            ws2['A1'].font = Font(bold=True, size=14, color="366092")
            ws2.merge_cells('A1:P1')
            
            # 写入表头
            option_df = pd.DataFrame(data['option_chain'])
            headers = list(option_df.columns)
            ws2.append(headers)
            
            for cell in ws2[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for _, row in option_df.iterrows():
                ws2.append(list(row))
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 15
            
            # 3. PCR数据sheet
            ws3 = wb.create_sheet("PCR数据")
            
            ws3['A1'] = "PCR（Put-Call Ratio）数据"
            ws3['A1'].font = Font(bold=True, size=14, color="366092")
            ws3.merge_cells('A1:I1')
            
            pcr_df = pd.DataFrame(data['pcr_data'])
            headers = list(pcr_df.columns)
            ws3.append(headers)
            
            for cell in ws3[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for _, row in pcr_df.iterrows():
                ws3.append(list(row))
            
            for col in range(1, len(headers) + 1):
                ws3.column_dimensions[get_column